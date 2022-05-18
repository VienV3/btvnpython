from distutils.file_util import write_file
import re
from shutil import ExecError
import requests
from openpyxl import load_workbook
from high_order_framework_requests_python_master import utils_class
string_interact1=utils_class.String_Interact()
    #thao tac voi file excel
def read_cell(file_path, sheetname, cellname):  #doc du lieu file excel
    wb=load_workbook(filename=file_path)
    sheet_ranges=wb[sheetname]
    return sheet_ranges[cellname].value
def update_cell(file_path,sheetname,cellname,newvalue):
    wb=load_workbook(filename=file_path)
    wb[sheetname][cellname].value=newvalue
    wb.close()
    wb.save(file_path)
#gom 2 phan:erorr,data
def get_direct_link(url_input):


    # url_input = 'https://www.youtube.com/watch?v=GxYpY-Wf9Ls'
    #step 1:get data fron ajax
    try:
        url = 'https://www.y2mate.com/mates/vi252/analyze/ajax'
        headers = {
            'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.36',
            'content-type':'application/x-www-form-urlencoded; charset=UTF-8'
        }
        dataPost = {
            "url":url_input,
            "q_auto":0,
            "ajax":1
        }
        res = requests.post(url,data=dataPost,headers=headers).text
        res= res.replace('\\','')
    except Exception as e:
        return '%s'%e,''
    #Step2: tach chuoi ra v_id, _id
    #utils_class.File_Interact('code.html').write_file(res)
    _id=string_interact1.regex_one_value('k__id = "(.*?)"',res) #su dung regex de lay duoc _id tu res.text
    v_id=string_interact1.regex_one_value('k_data_vid = "(.*?)"',res)#su dung regex de lay duoc v_id tu res.text
    # step3: 
    try:
        url="https://www.y2mate.com/mates/convert" #file url gianh de convert tu link can download sang link download
        headers = {
            'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.127 Safari/537.36',
            'content-type':'application/x-www-form-urlencoded; charset=UTF-8'
        }

        dataPOST={
            'type': 'youtube',
            '_id': _id,
            'v_id': v_id,
            'ajax': 1,
            'token': '',
            'ftype': 'mp4',
            'fquality': '720'
        } #thiet lap thong so cho link can download
        res = requests.post(url,data=dataPOST,headers=headers).text
        res=res.replace('\\','')
        utils_class.File_Interact('code.html').write_file(res)
        direct_link=string_interact1.regex_one_value('<a href="(.*?)"',res)
        return '',direct_link
    except Exception as e:
        return '%s'%e,''
if __name__=="__main__":
    # url_input='https://www.youtube.com/watch?v=GxYpY-Wf9Ls'
    # err,direct_link=get_direct_link(url_input)
    #bai toan ban dau, tai 1000.000 video
    
    L_input=[
        'https://www.youtube.com/watch?v=GxYpY-Wf9Ls',
        'https://www.youtube.com/watch?v=2b0-8qr2iGg',
        'https://www.youtube.com/watch?v=Q4r2p-9r1ds'
    ]
    L_output=[]
    file_path='ghifiledriectlink.xlsx'
    sheetname='Sheet1'
    update_cell(file_path,sheetname,'A1','STT')
    update_cell(file_path,sheetname,'B1','Link video')
    update_cell(file_path,sheetname,'C1','DirectLink')
    for url_input in range(0,len(L_input)):
        cell_link=L_input[url_input]
        direct_link=get_direct_link(cell_link)
        L_output.append(direct_link[1])
        cellname='B%s'%(url_input+2)
        update_cell(file_path,sheetname,cellname,cell_link) #update cell cot link video va tao ra link direct
    for url_output in range(0,len(L_output)):
        cell_directLink=L_output[url_input]
        cellname='C%s'%(url_output+2)
        update_cell(file_path,sheetname,cellname,cell_directLink)#update cell link direct
    
        #l_output =>save lai=>excel
        #url input, direct link=> duong dan toi file download
    


