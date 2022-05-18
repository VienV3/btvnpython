import requests
from openpyxl import load_workbook
from high_order_framework_requests_python_master import utils_class

string_Interact1=utils_class.String_Interact()
def read_cell(file_path,sheetname,cellname): #doc du lieu file excel
    wb=load_workbook(filename=file_path)
    sheet_ranges=wb[sheetname]
    return sheet_ranges[cellname].value
#update file cell
def update_cell(file_path,sheetname,cellname,newvalue):
    wb=load_workbook(filename=file_path)
    wb[sheetname][cellname].value=newvalue
    wb.close()
    wb.save(file_path)

if __name__=="__main__":
    l_link_post_total=[]
    l_title_post_total=[]
    file_path="ghifileweb.xlsx"
    sheetname="Sheet1"
    update_cell(file_path,sheetname,"A1","STT")
    update_cell(file_path,sheetname,"B1","Tên Bài Viết")
    update_cell(file_path,sheetname,"C1","Đường Dẫn Bài Viết")
    for pagenumber in range(1,10):
        url=f"http://nghiahsgs.com/page/{pagenumber}/"
        headers={
            "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36"
        }
        data=requests.get(url,headers=headers).text
        listCSSSelector=[
            '.entry-title a',
            '.entry-title',
        ]
        listeAttr=[
            'href',
            'innerText',
        ]
        l_link_post,l_title_post=string_Interact1.extractListTextByCSSSelector(data,listCSSSelector=listCSSSelector,listeAttr=listeAttr)
        l_title_post_total+=l_title_post
        l_link_post_total+=l_link_post
        if(l_title_post=="Oops! That page.. can’t be found."):
            break
        
        for postnumber in range(1,len(l_title_post_total)):
            title_post=l_title_post_total[postnumber]
            link_post=l_link_post_total[postnumber]
            cellname="A%s"%(postnumber+1)
            update_cell(file_path,sheetname,cellname,postnumber)

            cellname="B%s"%(postnumber+1)
            update_cell(file_path,sheetname,cellname,title_post)

            cellname="C%s"%(postnumber+1)
            update_cell(file_path,sheetname,cellname,link_post)