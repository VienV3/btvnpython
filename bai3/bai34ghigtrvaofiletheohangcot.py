from openpyxl import load_workbook

#doc file cel
def read_cell(file_path,sheetname,cellname):
    wb=load_workbook(filename=file_path)
    sheet_ranges=wb[sheetname]
    return sheet_ranges[cellname].value
#update file cell
def update_cell(file_path,sheetname,cellname,newvalue):
    wb=load_workbook(filename=file_path)
    wb[sheetname][cellname].value=newvalue
    wb.close()
    wb.save(file_path)

if __name__=="__main__":
    L1=[
        "Vũ Văn Viễn",
        "Đào Mạnh Hùng",
        "Võ Trọng Yến",
        "Hoàng Đức Hạnh",
        "Nguyễn Mạnh Hưng"
    ]
    L2=[30,31,24,12,23]
    #ghi file voi cot ho va ten, tuoi
    file_path="test.xlsx"
    sheetname="Sheet2"
    update_cell(file_path,sheetname,"A1","Họ Và Tên")
    update_cell(file_path,sheetname,"B1","Tuổi")
    for i_row in range(0,len(L1)):
        name=L1[i_row]
        age=L2[i_row]
        #print(name,age)
        cellname="A%s"%(i_row+2)
        update_cell(file_path,sheetname,cellname,name)
        cellname="B%s"%(i_row+2)
        update_cell(file_path,sheetname,cellname,age)
    #ghi file voi cot ho, ten dem, ten, tuoi
    file_path="test.xlsx"
    sheetname="Sheet3"
    update_cell(file_path,sheetname,"A1","Họ")
    update_cell(file_path,sheetname,"B1","Tên đệm")
    update_cell(file_path,sheetname,"C1","Tên")
    update_cell(file_path,sheetname,"D1","Tuổi")
    for i_row2 in range(0,len(L1)):
        name=L1[i_row2]
        age=L2[i_row2]
        ho,tendem,ten=name.split(" ")

        cellname="A%s"%(i_row2+2)
        update_cell(file_path,sheetname,cellname,ho)
        cellname="B%s"%(i_row2+2)
        update_cell(file_path,sheetname,cellname,tendem)
        cellname="C%s"%(i_row2+2)
        update_cell(file_path,sheetname,cellname,ten)
        cellname="D%s"%(i_row2+2)
        update_cell(file_path,sheetname,cellname,age)



