from openpyxl import load_workbook
#doc du lieu trong fiel excel 
def read_cell(file_path,sheetname,cell_name):
    wb=load_workbook(filename=file_path)
    sheet_ranges=wb[sheetname]
    return sheet_ranges[cell_name].value

def update_cell(file_path,sheetname,cellname,newvalue):
    wb=load_workbook(filename=file_path)
    wb[sheetname][cellname].value=newvalue
    wb.close()
    wb.save(file_path)

if __name__=="__main__":
    file_path="test.xlsx"
    sheetname="Sheet1"
    cellname="B1"
    #kq=read_cell(file_path,sheetname,cellname)
    #print(kq)
    newvalue="Đây là nội dung chỉnh sửa!"
    update_cell(file_path,sheetname,cellname,newvalue)












"""import openpyxl
def get_value_excel(filename,cellname):
    wb=openpyxl.load_workbook(filename)
    sheet1=wb["Sheet1"]
    wb.close()
    return sheet1[cellname].value
def update_value_excel(filename,cellname,value):
    wb=openpyxl.load_workbook(filename)
    sheet1=wb["Sheet1"]
    sheet1[cellname].value=value
    wb.close()
    wb.save(filename) """