from cv2 import split
import requests
from openpyxl import load_workbook
from high_order_framework_requests_python_master import utils_class
import re
String_Interact1=utils_class.String_Interact()
# btc= '52,597 BTC ($10,000,081,247) / +0.00284 BTC'
# Rice = re.search('(.*?) \(',btc).group(1)
# print(Rice)
# exit() cắt chuỗi với các kỹ tự đặc biệt
def read_cell(file_path,sheetname,cellname): #doc du lieu file excel
    wb=load_workbook(filename=file_path)
    sheet_ranges=wb[sheetname]
    return sheet_ranges[cellname].value
#update file cell
def update_cell(file_path,sheetname,cellname,newvalue):
    wb=load_workbook(filename=file_path)
    wb[sheetname][cellname].value=newvalue
    wb.close()
    wb.save(file_path)
if __name__=="__main__":
    L_current_wallet_total=[]
    file_path="vibtc.xlsx"
    sheetname="Sheet1"
    update_cell(file_path,sheetname,"A1","STT")
    update_cell(file_path,sheetname,"B1","Địa Chỉ Ví BTC")
    update_cell(file_path,sheetname,"C1","Giá Trị Ví")
    for pagenumber in range(1,2):
        
        url='https://bitinfocharts.com/top-100-richest-bitcoin-addresses.html'
        headers={
            "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.107 Safari/537.36"
        }
        data=requests.get(url,headers=headers).text
        listCSSSelector=[
            '.table-striped td>a',
            '.table-striped tbody tr td:nth-child(3)',
        ]
        listeAttr=[
            'innerText',
            'innerText',
        ]
        L_address_wallet,L_current_wallet=String_Interact1.extractListTextByCSSSelector(data,listCSSSelector=listCSSSelector,listeAttr=listeAttr)
        for tem_curent in range(0,len(L_current_wallet)):
            btc=L_current_wallet[tem_curent]
            cut_btc=re.search('(.*?) \(',btc).group(1)#cắt chuỗi tới 1 ký tự nào đó trong 1 chuỗi
            print(cut_btc)
        # for postnumber in range(1,len(L_address_wallet)):
            adress=L_address_wallet[tem_curent]
    # 
            cellname="A%s"%(tem_curent+2)
            update_cell(file_path,sheetname,cellname,tem_curent)
    # 
            cellname="B%s"%(tem_curent+2)
            update_cell(file_path,sheetname,cellname,adress)
    # 
            cellname="C%s"%(tem_curent+2)
            update_cell(file_path,sheetname,cellname,cut_btc)



